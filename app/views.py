import csv
import requests
import os
from django.conf import settings
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from rest_framework import status
from django.http import HttpResponse, JsonResponse
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from rest_framework.decorators import api_view
from .serializers import CSVUploadSerializer
from django.utils.timezone import now
from datetime import timedelta
from io import StringIO
from requests.auth import HTTPBasicAuth
from datetime import datetime
import json

LOGIN_URL = "https://api.baubuddy.de/index.php/login"
VEHICLE_API_URL = "https://api.baubuddy.de/dev/index.php/v1/vehicles/select/active"
LABEL_API_URL = "https://api.baubuddy.de/dev/index.php/v1/labels/{}"

UPLOAD_FOLDER = os.path.join(settings.BASE_DIR, 'app', 'upload')
# Color codes for conditional formatting
COLORS = {
    "green": "007500",
    "orange": "FFA500",
    "red": "b30000"
}
def generate_excel(response_data, keys, colored):
    """Generate an Excel file based on API response."""
    try:
        # Create DataFrame from response data
        df = pd.DataFrame(response_data)
                
        # Sort the DataFrame by 'gruppe'
        if "gruppe" in df.columns:
            df.sort_values(by="gruppe", inplace=True)
        
        # Add 'rnr' column with sequential values
        df["rnr"] = range(1, len(df) + 1)
        
        # Define the required columns
        keys_list = [key.strip() for key in keys.split(",")]
        for key in keys_list:
            df[key]=""
        columns = [col for col in df.columns if col not in keys_list]+[key for key in keys_list]
        
        csv_df= df[columns]
        
         # Step 2: Get the access token for authentication
        access_token = get_access_token()
        
        # Step 3: Fetch vehicle data from the external API
        vehicle_data = get_vehicle_data(access_token)

        # Step 4: Filter out vehicles that do not have the 'hu' field
        filtered_vehicles = [vehicle for vehicle in vehicle_data if vehicle.get('hu')]
        
       
        # Step 5: covert vehicle data to dataframe
        
        vehicle_df = pd.DataFrame(vehicle_data)
       
                    
        # merged the csv_data & vehicle_data
        merged_df = pd.merge(csv_df, vehicle_df, on="gruppe", how="outer",suffixes=("_csv", "_api"))

        # get the colorcodes by labelIDs 
        merged_df['colorCodes'] = merged_df.apply(
                lambda row: resolve_combined_label_colors(row, access_token) 
                if (pd.notna(row['labelIds_csv']) and row['labelIds_csv'] != "") or 
                   (pd.notna(row['labelIds_api']) and row['labelIds_api'] != "")
                else [],  # Return an empty list if both labelIds are empty or NaN
                axis=1) 
        
        # Create Excel file
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Vehicle Data"

        headers = merged_df.columns.tolist()
        sheet.append(headers)
       
        # Fill in data
        for _, row in merged_df.iterrows(): 
            
            row_values = [str(value) if isinstance(value, list) else value for value in row.values]
            sheet.append(row_values)
           
            if colored and (
                ("hu_csv" in row or "hu_api" in row) and 
                (pd.notna(row["hu_csv"]) or pd.notna(row["hu_api"]) and 
                ((row["hu_csv"] or row["hu_api"]) != ""))):
                try:
                    hu_value = row["hu_csv"] if pd.notna(row["hu_csv"]) and row["hu_csv"] != "" else row["hu_api"]
                    
                    hu_date = pd.to_datetime(hu_value, errors="coerce")
                    if pd.notna(hu_date):
                        if hu_date.tzinfo is not None:
                            hu_date = hu_date.tz_localize(None)

                        now = datetime.now()
                        delta_days = (now - hu_date).days
                        

                        if delta_days <= 90:
                            color_fill = PatternFill(start_color=COLORS["green"], end_color=COLORS["green"], fill_type="solid")
                        elif delta_days <= 365:
                            color_fill = PatternFill(start_color=COLORS["orange"], end_color=COLORS["orange"], fill_type="solid")
                        else:
                            color_fill = PatternFill(start_color=COLORS["red"], end_color=COLORS["red"], fill_type="solid")

                        for col_idx in range(1, len(row_values) + 1):
                            sheet.cell(row=sheet.max_row, column=col_idx).fill = color_fill

                except Exception as e:
                    print(f"Error processing 'hu' date for row: {e}")

        # Return the Excel file as a byte stream
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        return excel_file
   
    except UnicodeDecodeError as ude:
        print(f"Unicode error encountered: {ude}")
        raise

    except Exception as e:
        print(f"Error generating Excel file: {e}")
        raise


@api_view(['Post'])
def upload_csv(request):
    """Handle CSV upload and Excel generation via the REST API."""
    serializer = CSVUploadSerializer(data=request.data)
   
    if serializer.is_valid():
        
        # Get the uploaded file and data
        csv_file = serializer.validated_data.get('csv_file')
        keys = serializer.validated_data.get("keys", "")
        colored = serializer.validated_data.get("colored", True)
                  
        
        # Check if the file is a Django InMemoryUploadedFile or TemporaryUploadedFile
        if hasattr(csv_file, 'read'):
            file_content =csv_file.read().decode('utf-8').strip() 
            file_content = file_content.replace("\r\n", "\n")
            file_content = "\n".join(line for line in file_content.splitlines() if line.strip())
                
            file = StringIO(file_content)  
    
        df = pd.read_csv(file,delimiter=';')
        csv_data = df.to_json(orient='records')
        csv_data=json.loads(csv_data)
          

        # Generate Excel file
        excel_file=generate_excel(csv_data, keys, colored)
        current_date=now().isoformat()[:10]
         # Set up the response to return the Excel file
        response =HttpResponse(
            excel_file.read(),  # Read the file contents
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
       
        # Set content disposition to prompt a file download
        response["Content-Disposition"] = f'attachment; filename="vehicles_{current_date}.xlsx"'
        return response
        

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Function to get the access token
def get_access_token():
    payload = {
        "username": "365",
        "password": "1"
    }
    headers = {
        "Authorization": "Basic QVBJX0V4cGxvcmVyOjEyMzQ1NmlzQUxhbWVQYXNz",
        "Content-Type": "application/json"
    }
    response = requests.post(LOGIN_URL, json=payload, headers=headers)
    if response.status_code == 200:
        return response.json().get("oauth", {}).get("access_token")
    else:
        raise Exception("Failed to get access token")


# Function to get vehicle data from the external API
def get_vehicle_data(access_token):
    headers = {
        "Authorization": f"Bearer {access_token}"
    }
    response = requests.get(VEHICLE_API_URL, headers=headers)
    if response.status_code == 200:
        return response.json()
    else:
        raise Exception("Failed to fetch vehicle data")


# Function to get label data from the external API
def get_label_data(label_id, access_token):
    headers = {
        "Authorization": f"Bearer {access_token}"
    }
    url = LABEL_API_URL.format(label_id)
   
    response = requests.get(f"{LABEL_API_URL}", headers=headers)
    if response.status_code == 200:
       return response.json()
    else:
        return ""
       


# Define a function to merge and resolve labelIds
def resolve_combined_label_colors(row, access_token):
    """
    Resolve colorCodes for combined labelIds from CSV and API sources.

    Parameters:
    - row: A row of the DataFrame.
    - access_token: API access token.

    Returns:
    - List of resolved colorCodes.
    """
    # Extract labelIds from both sources
    label_ids_csv = str(row.get("labelIds_csv", "")) if pd.notna(row.get("labelIds_csv")) else ""
    label_ids_api = str(row.get("labelIds_api", "")) if pd.notna(row.get("labelIds_api")) else ""

    # Combine labelIds
    combined_label_ids = []
    if pd.notna(label_ids_csv) and label_ids_csv != "":
        combined_label_ids.extend(label_ids_csv.split(','))
    if pd.notna(label_ids_api) and label_ids_api != "":
        combined_label_ids.extend(label_ids_api.split(','))

    # Remove duplicates
    combined_label_ids = list(set(map(str.strip, combined_label_ids)))

    # Resolve colorCodes for each labelId
    color_codes = [
        get_label_data(int(label_id), access_token)
        for label_id in combined_label_ids
        if label_id.isdigit()  # Ensure the labelId is a valid integer
    ]
    return color_codes








   